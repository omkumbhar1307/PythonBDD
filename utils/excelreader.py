import os
import openpyxl

class ExcelReader:
    def __init__(self, file_name, sheet_name=None):
        """Initialize with file path and sheet name (optional)."""

        # Get the absolute path of the project root directory
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))  # utils/
        ROOT_DIR = os.path.dirname(ROOT_DIR)  # Moves up to ui_testing_project/

        # Construct full path dynamically
        self.file_path = os.path.join(ROOT_DIR, "testdata", file_name)

        # Check if the file exists before loading
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Error: Excel file NOT found at {self.file_path}")

        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active

    def get_cell_value(self, row, col):
        """Get value from a specific cell (1-based index)."""
        cellValue= self.sheet.cell(row=row, column=col).value
        ExcelReader.close(self)
        return cellValue

    def get_row_values(self, row):
        """Get all values from a specific row (1-based index)."""
        rowValues = [cell.value for cell in self.sheet[row]]
        ExcelReader.close(self)
        return rowValues

    def get_column_values(self, col):
        """Get all values from a specific column (1-based index)."""
        columnValues = [self.sheet.cell(row=i, column=col).value for i in range(1, self.sheet.max_row + 1)]
        ExcelReader.close(self)
        return columnValues

    def get_all_data(self):
        """Get all data from the sheet as a list of lists."""
        allData = [[cell.value for cell in row] for row in self.sheet.iter_rows()]
        ExcelReader.close(self)
        return allData

    def close(self):
        """Close the workbook."""
        self.workbook.close()
